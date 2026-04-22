#!/usr/bin/env python
"""
Convert knee metadata Excel workbook to JSON.

This script reads the knee metadata Excel workbook and generates a JSON file.
Data is validated against the rules defined in src.config.knee_metadata.

Usage:
    excel-to-json [input_excel] [output_json]

Arguments:
    input_excel   Path to input Excel file (default: data/knee_metadata.xlsx)
    output_json   Path to output JSON file (default: src/config/metadata/knee_metadata.json)
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Dict, List, Any, Tuple, Optional, Union

VideoId = Union[str, int]

# Add project src/ to sys.path so we can import config.knee_metadata
PROJECT_ROOT = Path(__file__).resolve().parents[3]
SRC_PATH = PROJECT_ROOT / "src"
if str(SRC_PATH) not in sys.path:
    sys.path.insert(0, str(SRC_PATH))

try:
    from config.knee_metadata import validate_knee_metadata, KneeVideoMeta, Cycle, FrameRange, RegionSegments
except ImportError as e:
    sys.stderr.write(f"Error: Could not import validation module: {e}\n")
    sys.stderr.write("Ensure the knee-segmentation package is installed or src/ is on PYTHONPATH.\n")
    sys.exit(1)


def parse_worksheet_name(name: str) -> Optional[Tuple[VideoId, int]]:
    """
    Parse worksheet name to extract video_id and n_segments.
    
    Returns (video_id, n_segments) where video_id is kept as a string to preserve
    any leading zeros (e.g., "0000"). Returns None if parsing fails.
    """
    match = re.match(r"^(.+?)_N(\d+)$", name)
    if not match:
        return None
    video_id_str = match.group(1)
    n_segments = int(match.group(2))
    return (video_id_str, n_segments)


def read_excel_to_data(input_path: Path) -> Tuple[Dict[str, Any], List[str]]:
    """
    Read Excel workbook and convert to JSON data structure.
    
    Returns (data_dict, errors) where data_dict has keys schema_version and videos.
    """
    from openpyxl import load_workbook
    
    errors: List[str] = []
    videos: List[Dict[str, Any]] = []
    
    try:
        wb = load_workbook(input_path, read_only=True)
    except Exception as e:
        return {"schema_version": 1, "videos": []}, [f"Failed to open Excel file: {e}"]
    
    # Validate metadata sheet
    if "_metadata" not in wb.sheetnames:
        errors.append("Workbook missing required '_metadata' sheet")
    else:
        ws_meta = wb["_metadata"]
        if ws_meta["B1"].value != 1:
            errors.append(f"Metadata sheet: schema_version={ws_meta['B1'].value!r}, expected 1")
    
    # Process video worksheets
    video_sheets = [s for s in wb.sheetnames if s != "_metadata"]
    
    for sheet_name in video_sheets:
        ws = wb[sheet_name]
        ws.reset_dimensions()
        ctx = f"worksheet [{sheet_name}]"
        
        # Parse name
        parsed = parse_worksheet_name(sheet_name)
        if parsed is None:
            errors.append(f"{ctx}: invalid worksheet name format (expected '<video_id>_N<n_segments>')")
            continue
        video_id_raw, n_segments = parsed
        
        # Read metadata cells
        try:
            meta_video_id = ws["B1"].value
            meta_n_segments = ws["B2"].value
            condition = ws["B3"].value
        except Exception as e:
            errors.append(f"{ctx}: failed to read metadata cells: {e}")
            continue
        
        # Validate metadata consistency
        if str(meta_video_id) != str(video_id_raw):
            errors.append(f"{ctx}: video_id mismatch; name says {video_id_raw!r}, cell says {meta_video_id!r}")
        if meta_n_segments != n_segments:
            errors.append(f"{ctx}: n_segments mismatch; name says {n_segments}, cell says {meta_n_segments}")
        if not condition:
            errors.append(f"{ctx}: missing condition")
        
        # Read regions (rows 5-8)
        regions: Dict[str, Dict[str, int]] = {}
        for idx, region_name in enumerate(["JC", "OT", "SB"]):
            row = 6 + idx
            try:
                s1 = ws[f"B{row}"].value
                e1 = ws[f"C{row}"].value
                if s1 is None or e1 is None:
                    errors.append(f"{ctx}: region {region_name} missing start or end value")
                    continue
                regions[region_name] = {"start_1": int(s1), "end_1": int(e1)}
            except Exception as e:
                errors.append(f"{ctx}: error reading region {region_name}: {e}")
        
        # Read cycles (starting row 10)
        cycles: List[Dict[str, Any]] = []
        row = 11
        while ws[f"A{row}"].value is not None:
            try:
                flex_start = ws[f"B{row}"].value
                flex_end = ws[f"C{row}"].value
                ext_start = ws[f"D{row}"].value
                ext_end = ws[f"E{row}"].value
                if None in (flex_start, flex_end, ext_start, ext_end):
                    break
                cycles.append({
                    "flex": {"start_1": int(flex_start), "end_1": int(flex_end)},
                    "ext": {"start_1": int(ext_start), "end_1": int(ext_end)}
                })
            except Exception as e:
                errors.append(f"{ctx}: error reading cycle at row {row}: {e}")
            row += 1
        
        # Append video entry (preserve raw video_id)
        videos.append({
            "video_id": video_id_raw,
            "n_segments": n_segments,
            "condition": condition,
            "regions": regions,
            "cycles": cycles
        })
    
    wb.close()
    
    # Build validation registry (convert video_id to int for keys)
    registry: Dict[Tuple[int, int], KneeVideoMeta] = {}
    for entry in videos:
        video_id_raw = entry["video_id"]
        n_segments = entry["n_segments"]
        try:
            video_id_int = int(video_id_raw)
        except (ValueError, TypeError):
            errors.append(f"video_id {video_id_raw!r} (from worksheet name) is not numeric")
            continue
        
        try:
            # Build RegionSegments
            regions_obj: Dict[str, RegionSegments] = {}
            for name in ("JC", "OT", "SB"):
                r = entry["regions"].get(name)
                if r is None:
                    raise KeyError(f"missing region {name}")
                regions_obj[name] = RegionSegments(s=r["start_1"], e=r["end_1"])
            
            # Build cycles
            cycles_list: List[Cycle] = []
            for c in entry["cycles"]:
                flex = FrameRange.from_1based(c["flex"]["start_1"], c["flex"]["end_1"])
                ext = FrameRange.from_1based(c["ext"]["start_1"], c["ext"]["end_1"])
                cycles_list.append(Cycle(flex=flex, ext=ext))
            
            meta = KneeVideoMeta(
                condition=entry["condition"] or "",
                video_id=video_id_int,
                n_segments=n_segments,
                cycles=cycles_list,
                regions=regions_obj
            )
            key = (video_id_int, n_segments)
            if key in registry:
                errors.append(f"Duplicate key {key} (video_id={video_id_raw}, n_segments={n_segments})")
            else:
                registry[key] = meta
        except Exception as e:
            errors.append(f"Failed to build metadata for video {video_id_raw} (N{n_segments}): {e}")
    
    # Run official validation
    if not errors:
        try:
            validate_knee_metadata(registry)
        except ValueError as e:
            errors.append(str(e))
    
    return {"schema_version": 1, "videos": videos}, errors


def main():
    parser = argparse.ArgumentParser(
        description="Convert knee metadata Excel to JSON with validation"
    )
    parser.add_argument(
        "input_excel",
        type=str,
        nargs="?",
        default="data/knee_metadata.xlsx",
        help="Path to input Excel file (default: data/knee_metadata.xlsx)"
    )
    parser.add_argument(
        "output_json",
        type=str,
        nargs="?",
        default="src/config/metadata/knee_metadata.json",
        help="Path to output JSON file (default: src/config/metadata/knee_metadata.json)"
    )
    
    args = parser.parse_args()
    
    input_path = Path(args.input_excel)
    output_path = Path(args.output_json)
    
    if not input_path.exists():
        print(f"Error: Input Excel file not found: {input_path}", file=sys.stderr)
        sys.exit(1)
    
    data, errors = read_excel_to_data(input_path)
    
    if errors:
        print("Validation errors found:")
        for error in errors:
            print(f"  - {error}")
        sys.exit(1)
    
    # Write JSON
    output_path.parent.mkdir(parents=True, exist_ok=True)
    json_str = json.dumps(data, indent=2)
    output_path.write_text(json_str, encoding="utf-8")
    print(f"JSON written to: {output_path}")


if __name__ == "__main__":
    main()
