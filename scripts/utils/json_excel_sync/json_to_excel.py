#!/usr/bin/env python
"""
Convert knee metadata JSON to Excel workbook.

This script reads the knee metadata JSON file and generates an Excel workbook
with one worksheet per (video_id, n_segments) combination.

Usage:
    json_to_excel.py [input_json] [output_excel]

Arguments:
    input_json   Path to input JSON file (default: src/config/metadata/knee_metadata.json)
    output_excel Path to output Excel file (default: data/knee_metadata.xlsx)
"""

import argparse
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Tuple


Key = Tuple[int, int]  # (video_id, n_segments)


def load_knee_metadata_json(json_path: Path) -> Dict[str, Any]:
    """Load and parse the knee metadata JSON file."""
    if not json_path.exists():
        raise FileNotFoundError(f"JSON file not found: {json_path}")
    
    raw = json.loads(json_path.read_text(encoding="utf-8"))
    
    if not isinstance(raw, dict):
        raise ValueError(f"JSON root must be an object; got {type(raw).__name__}")
    
    schema_version = raw.get("schema_version", None)
    if schema_version != 1:
        raise ValueError(f"Unsupported schema_version={schema_version!r}; expected 1")
    
    videos = raw.get("videos", None)
    if not isinstance(videos, list):
        raise ValueError(f"JSON 'videos' must be a list; got {type(videos).__name__}")
    
    return raw


def group_videos_by_key(videos: List[Dict[str, Any]]) -> Dict[Tuple[Any, int], List[Dict[str, Any]]]:
    """Group video entries by (video_id, n_segments) key, preserving video_id type."""
    groups: Dict[Tuple[Any, int], List[Dict[str, Any]]] = {}
    
    for entry in videos:
        # Preserve video_id as-is (could be int or string)
        video_id = entry["video_id"]
        n_segments = int(entry["n_segments"])
        key = (video_id, n_segments)
        
        if key not in groups:
            groups[key] = []
        groups[key].append(entry)
    
    return groups


def create_worksheet_name(video_id, n_segments: int) -> str:
    """Create worksheet name from video_id and n_segments.
    
    Args:
        video_id: Can be int or str - will be converted to string as-is
        n_segments: Number of segments
    
    Returns:
        Worksheet name in format {video_id}_N{n_segments}
    """
    return f"{video_id}_N{n_segments}"


def write_excel_from_json(input_path: Path, output_path: Path) -> None:
    """Convert JSON metadata to Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    # Bold font for labels
    bold_font = Font(bold=True)
    
    # Load JSON data
    data = load_knee_metadata_json(input_path)
    videos = data["videos"]
    
    # Group by (video_id, n_segments)
    groups = group_videos_by_key(videos)
    
    # Create workbook
    wb = Workbook()
    ws_metadata = wb.active
    ws_metadata.title = "_metadata"
    
    # Write metadata sheet
    ws_metadata["A1"] = "schema_version"
    ws_metadata["A1"].font = bold_font
    ws_metadata["B1"] = 1
    ws_metadata["A2"] = "last_modified"
    ws_metadata["A2"].font = bold_font
    ws_metadata["B2"] = datetime.now().isoformat()
    ws_metadata["A3"] = "source"
    ws_metadata["A3"].font = bold_font
    ws_metadata["B3"] = "Excel (source of truth)"
    
    # Create worksheets for each video group
    # Sort by n_segments then by string representation of video_id (handles mixed int/str)
    sorted_groups = sorted(groups.items(), key=lambda kv: (kv[0][1], str(kv[0][0])))
    
    for i, (key, entries) in enumerate(sorted_groups, start=1):
        video_id, n_segments = key
        
        # Create new worksheet - use str(video_id) to ensure proper formatting
        ws_name = create_worksheet_name(str(video_id), n_segments)
        
        # Ensure worksheet name is unique (Excel max 31 chars, but we should be fine)
        if ws_name in wb.sheetnames:
            ws_name = f"{ws_name}_{i}"
        
        ws = wb.create_sheet(title=ws_name)
        
        # Write video metadata (row 1)
        entry = entries[0]  # All entries in a group should be identical
        
        ws["A1"] = "video_id"
        ws["A1"].font = bold_font
        ws["B1"] = video_id
        ws["A2"] = "n_segments"
        ws["A2"].font = bold_font
        ws["B2"] = n_segments
        ws["A3"] = "condition"
        ws["A3"].font = bold_font
        ws["B3"] = entry["condition"]
        
        # Write regions (row 5)
        ws["A5"] = "Region"
        ws["A5"].font = bold_font
        ws["B5"] = "start_1"
        ws["B5"].font = bold_font
        ws["C5"] = "end_1"
        ws["C5"].font = bold_font
        
        regions = entry["regions"]
        for row_idx, region_name in enumerate(["JC", "OT", "SB"], start=6):
            ws[f"A{row_idx}"] = region_name
            ws[f"A{row_idx}"].font = bold_font
            ws[f"B{row_idx}"] = regions[region_name]["start_1"]
            ws[f"C{row_idx}"] = regions[region_name]["end_1"]
        
        # Write cycles (row 10)
        ws["A10"] = "cycle_index"
        ws["A10"].font = bold_font
        ws["B10"] = "flex_start_1"
        ws["B10"].font = bold_font
        ws["C10"] = "flex_end_1"
        ws["C10"].font = bold_font
        ws["D10"] = "ext_start_1"
        ws["D10"].font = bold_font
        ws["E10"] = "ext_end_1"
        ws["E10"].font = bold_font
        
        for cycle_idx, cycle in enumerate(entry["cycles"], start=11):
            ws[f"A{cycle_idx}"] = cycle_idx - 10
            ws[f"B{cycle_idx}"] = cycle["flex"]["start_1"]
            ws[f"C{cycle_idx}"] = cycle["flex"]["end_1"]
            ws[f"D{cycle_idx}"] = cycle["ext"]["start_1"]
            ws[f"E{cycle_idx}"] = cycle["ext"]["end_1"]
    
    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Excel workbook written to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert knee metadata JSON to Excel workbook"
    )
    parser.add_argument(
        "input_json",
        type=str,
        nargs="?",
        default="src/config/metadata/knee_metadata.json",
        help="Path to input JSON file (default: src/config/metadata/knee_metadata.json)"
    )
    parser.add_argument(
        "output_excel",
        type=str,
        nargs="?",
        default="data/knee_metadata.xlsx",
        help="Path to output Excel file (default: data/knee_metadata.xlsx)"
    )
    
    args = parser.parse_args()
    
    input_path = Path(args.input_json)
    output_path = Path(args.output_excel)
    
    try:
        write_excel_from_json(input_path, output_path)
    except FileNotFoundError as e:
        print(f"Error: {e}")
        exit(1)
    except ValueError as e:
        print(f"Validation error: {e}")
        exit(1)
    except Exception as e:
        print(f"Unexpected error: {e}")
        exit(1)


if __name__ == "__main__":
    main()
